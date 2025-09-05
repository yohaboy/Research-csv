from celery import chain

from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework import status, parsers
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated ,AllowAny

from reportlab.lib.pagesizes import letter,A4
from reportlab.pdfgen import canvas

from django.db.models import Q
from datetime import datetime
from .models import Author, ResearchGroup, Publication, AuthorPublication
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from .models import Author, ResearchGroup, Publication, AuthorPublication
import openpyxl
from openpyxl.utils import get_column_letter
from .serializers import (
    AuthorSerializer,
    ResearchGroupSerializer,
    PublicationSerializer,
    AuthorPublicationSerializer
)

from django_q.tasks import async_task
from django_q.models import Task


class BaseAPIView(APIView):
    permission_classes = [AllowAny]

    def get_since_date(self, request):
        since = request.query_params.get("since")
        if since:
            try:
                return datetime.strptime(since, "%Y-%m-%d").date()
            except ValueError:
                pass
        return None


class IndexAPIView(APIView):
    def get(self, request):
        data = {
            "authors_count": Author.objects.count(),
            "publications_count": Publication.objects.count(),
            "research_groups_count": ResearchGroup.objects.count(),
            "author_publications_count": AuthorPublication.objects.count(),
        }
        return Response(data)


class AuthorList(generics.ListAPIView):
    queryset = Author.objects.select_related("research_group").all()
    serializer_class = AuthorSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.query_params.get("search", "")
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query)
                | Q(last_name__icontains=search_query)
            )
        return queryset


class ResearchGroupList(generics.ListAPIView):
    queryset = ResearchGroup.objects.all()
    serializer_class = ResearchGroupSerializer
    permission_classes = [AllowAny]


class PublicationList(APIView):
    def get(self, request, *args, **kwargs):
        selected_group = request.GET.get("group")
        selected_author = request.GET.get("author")
        selected_source = request.GET.get("source")

        publications = Publication.objects.all()

        if selected_group:
            publications = publications.filter(
                authorpublication__author__research_group__id=selected_group
            )
        if selected_author:
            publications = publications.filter(
                authorpublication__author__id=selected_author
            )
        if selected_source:
            publications = publications.filter(source=selected_source)

        publications = publications.distinct()

        pub_serializer = PublicationSerializer(publications, many=True)
        grp_serializer = ResearchGroupSerializer(
            ResearchGroup.objects.all(), many=True
        )
        author_serializer = AuthorSerializer(Author.objects.all(), many=True)

        return Response(
            {
                "publications": pub_serializer.data,
                "research_groups": grp_serializer.data,
                "authors": author_serializer.data,
            }
        )


class FileUploadView(BaseAPIView):
    parser_classes = [
        parsers.MultiPartParser,
        parsers.FormParser,
        parsers.JSONParser,
    ]

    def post(self, request, *args, **kwargs):
        if request.data.get("clear_data"):
            AuthorPublication.objects.all().delete()
            Publication.objects.all().delete()
            Author.objects.all().delete()
            ResearchGroup.objects.all().delete()
            return Response(
                {"message": "All data cleared."}, status=status.HTTP_200_OK
            )

        uploaded_file = request.FILES.get("csv_file")
        if not uploaded_file:
            return Response(
                {"error": "No file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = uploaded_file.name.lower()
        if not filename.endswith((".csv", ".xls", ".xlsx")):
            return Response(
                {"error": "File must be .csv, .xls, or .xlsx format."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            file_bytes = uploaded_file.read()
            from .tasks import process_file_upload, fetch_publications_task

            task_id = async_task(
                "research.tasks.process_file_upload",
                file_bytes,
                filename,
                hook="research.tasks.fetch_publications_task",  # callback after upload
            )

            return Response(
                {"message": "File upload started.", "task_id": task_id},
                status=status.HTTP_202_ACCEPTED,
            )
        except Exception as e:
            return Response(
                {"error": f"Error processing file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class NewPublicationsCount(BaseAPIView):
    def get(self, request):
        since_date = self.get_since_date(request)
        if not since_date:
            return Response(
                {
                    "error": "Invalid or missing 'since' parameter (format: YYYY-MM-DD)"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        papers = Publication.objects.filter(publication_date__gt=since_date)
        serialized_papers = PublicationSerializer(papers, many=True)
        return Response(
            {"count": papers.count(), "since": since_date, "papers": serialized_papers.data}
        )


class KeywordCounts(BaseAPIView):
    def get(self, request):
        since_date = self.get_since_date(request)
        qs = Publication.objects.all()
        if since_date:
            qs = qs.filter(publication_date__gt=since_date)

        keyword_counter = {}
        for pub in qs:
            for kw in pub.keywords.split(","):
                kw = kw.strip().lower()
                if kw:
                    keyword_counter[kw] = keyword_counter.get(kw, 0) + 1

        return Response({"keywords": keyword_counter, "since": since_date})


class MultiGroupPapersCount(BaseAPIView):
    def get(self, request):
        multi_group_pub_ids = set()
        for pub in Publication.objects.all():
            groups = set(
                pub.authorpublication_set.select_related(
                    "author__research_group"
                ).values_list("author__research_group__name", flat=True)
            )
            if len(groups) > 1:
                multi_group_pub_ids.add(pub.id)

        multi_group_pubs_qs = Publication.objects.filter(id__in=multi_group_pub_ids)
        serialized_pubs = PublicationSerializer(multi_group_pubs_qs, many=True)

        return Response(
            {"count": multi_group_pubs_qs.count(), "publications": serialized_pubs.data}
        )


class GroupAuthorMultiGroup(BaseAPIView):
    def get(self, request):
        result = {}
        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)
            author_counts = {}
            for author in authors:
                count = 0
                for ap in author.authorpublication_set.all():
                    pub = ap.publication
                    groups = set(
                        pub.authorpublication_set.select_related(
                            "author__research_group"
                        ).values_list("author__research_group__name", flat=True)
                    )
                    if len(groups) > 1:
                        count += 1
                if count > 0:
                    author_counts[str(author)] = count
            if author_counts:
                result[group.name] = author_counts
        return Response({"data": result})


class TotalPapersPerGroup(BaseAPIView):
    def get(self, request):
        result = {}
        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)
            pub_ids = set()
            for author in authors:
                pub_ids.update(
                    author.authorpublication_set.values_list("publication_id", flat=True)
                )
            pubs = Publication.objects.filter(id__in=pub_ids)
            serializer = PublicationSerializer(pubs, many=True)
            result[group.name] = serializer.data
        return Response({"data": result})


class KeywordCountsPerGroup(BaseAPIView):
    def get(self, request):
        result = {}
        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)
            pub_ids = set()
            for author in authors:
                pub_ids.update(
                    author.authorpublication_set.values_list("publication_id", flat=True)
                )
            keyword_counter = {}
            pubs = Publication.objects.filter(id__in=pub_ids)
            for pub in pubs:
                for kw in pub.keywords.split(","):
                    kw = kw.strip().lower()
                    if kw:
                        keyword_counter[kw] = keyword_counter.get(kw, 0) + 1
            result[group.name] = keyword_counter
        return Response(result)


class TriggerFetchPublications(BaseAPIView):
    def post(self, request):
        from .tasks import fetch_publications_task
        task_id = async_task("research.tasks.fetch_publications_task")
        return Response({"status": "started", "task_id": task_id})


class CheckTaskStatus(BaseAPIView):
    pass
    # def get(self, request):
    #     task_id = request.query_params.get("task_id")
    #     if not task_id:
    #         return Response(
    #             {"error": "No task_id provided"}, status=status.HTTP_400_BAD_REQUEST
    #         )

    #     try:
    #         task = Task.objects.get(id=task_id)
    #         return Response(
    #             {
    #                 "task_id": str(task.id),
    #                 "success": task.success,
    #                 "result": task.result,
    #                 "started": task.started,
    #                 "stopped": task.stopped,
    #             }
    #         )
    #     except Task.DoesNotExist:
    #         return Response({"error": "Task not found"}, status=status.HTTP_404_NOT_FOUND)
        
   
class ExportAllExcel(APIView):
    def get(self, request):
        def strip_tz(value):
            return value.replace(tzinfo=None) if hasattr(value, 'tzinfo') else value

        wb = openpyxl.Workbook()

        # Sheet 1: Research Groups
        ws1 = wb.active
        ws1.title = "Research Groups"
        ws1.append(["ID", "Name"])
        research_groups = ResearchGroup.objects.all()
        for group in research_groups:
            ws1.append([group.id, group.name])

        # Sheet 2: Authors
        ws2 = wb.create_sheet("Authors")
        ws2.append([
            "ID", "First Name", "Last Name", "Group", "Scopus ID",
            "Scholar ID", "ORCID ID", "Staff URL"
        ])
        authors = Author.objects.select_related('research_group').all()
        for author in authors:
            ws2.append([
                author.id, author.first_name, author.last_name,
                author.research_group.name if author.research_group else "N/A",
                author.scopus_id, author.scholar_id, author.orcid_id, author.staff_url
            ])

        # Sheet 3: Publications
        ws3 = wb.create_sheet("Publications")
        ws3.append([
            "ID", "Title", "Publication Date", "Keywords", "Abstract", "Date Added", "URL", "Source"
        ])
        publications = Publication.objects.all()
        for pub in publications:
            ws3.append([
                pub.id,
                pub.title,
                pub.publication_date,
                pub.keywords,
                pub.abstract,
                strip_tz(pub.date_added) if pub.date_added else '',
                pub.url,
                pub.source
            ])

        # Sheet 4: Author-Publication
        ws4 = wb.create_sheet("Author-Publication")
        ws4.append(["ID", "Author", "Publication", "Order"])
        links = AuthorPublication.objects.select_related('author', 'publication').all()
        for ap in links:
            ws4.append([
                ap.id, str(ap.author), ap.publication.title, ap.author_order
            ])

        # Auto-fit column width
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(10, min(max_length + 2, 50))

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=all_data.xlsx'
        wb.save(response)
        return response


class ExportAllPDF(APIView):
    def get(self, request):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        styles = getSampleStyleSheet()
        elements = []

        # Research Groups
        elements.append(Paragraph("Research Groups", styles['Title']))
        for group in ResearchGroup.objects.all():
            elements.append(Paragraph(f"- {group.name}", styles['Normal']))
        elements.append(PageBreak())

        # Authors
        elements.append(Paragraph("Authors", styles['Title']))
        for author in Author.objects.select_related('research_group'):
            elements.append(Paragraph(
                f"{author.first_name} {author.last_name} | Group: {author.research_group.name} | "
                f"Scopus: {author.scopus_id or 'N/A'} | Scholar: {author.scholar_id or 'N/A'} | ORCID: {author.orcid_id or 'N/A'}",
                styles['Normal']
            ))
        elements.append(PageBreak())

        # Publications
        elements.append(Paragraph("Publications", styles['Title']))
        for pub in Publication.objects.all():
            elements.append(Paragraph(f"Title: {pub.title}", styles['Heading3']))
            elements.append(Paragraph(f"Date: {pub.publication_date}", styles['Normal']))
            elements.append(Paragraph(f"Source: {pub.source}", styles['Normal']))
            elements.append(Paragraph(f"URL: {pub.url}", styles['Normal']))
            elements.append(Paragraph(f"Keywords: {pub.keywords}", styles['Normal']))
            elements.append(Paragraph(f"Abstract: {pub.abstract}", styles['Normal']))
            elements.append(Spacer(1, 12))
        elements.append(PageBreak())

        # Author-Publication Links
        elements.append(Paragraph("Author-Publication Links", styles['Title']))
        for ap in AuthorPublication.objects.select_related('author', 'publication'):
            elements.append(Paragraph(
                f"Author: {ap.author} | Publication: {ap.publication.title} | Order: {ap.author_order}",
                styles['Normal']
            ))

        doc.build(elements)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')


class ExportFilteredExcel(APIView):
    def get(self, request):
        since = request.query_params.get('since')
        try:
            since_date = datetime.strptime(since, '%Y-%m-%d').date()
        except Exception:
            return Response({'error': 'Invalid date format, use YYYY-MM-DD'}, status=400)

        publications = Publication.objects.filter(publication_date__gt=since_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filtered Publications"
        ws.append(["ID", "Title", "Publication Date", "Keywords", "Abstract", "URL", "Source"])

        for pub in publications:
            ws.append([
                pub.id, pub.title, pub.publication_date.strftime('%Y-%m-%d'),
                pub.keywords, pub.abstract, pub.url, pub.source
            ])

        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 50))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=filtered_publications.xlsx'
        wb.save(response)
        return response



class ExportFilteredPDF(APIView):
    def get(self, request):
        since = request.query_params.get('since')
        try:
            since_date = datetime.strptime(since, '%Y-%m-%d').date()
        except Exception:
            return Response({'error': 'Invalid date format, use YYYY-MM-DD'}, status=400)

        publications = Publication.objects.filter(publication_date__gt=since_date)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=filtered_publications.pdf'

        c = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        y = height - 40
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, y, "Filtered Publications Since " + since)
        c.setFont("Helvetica", 12)
        y -= 30

        for pub in publications:
            if y < 50:
                c.showPage()
                y = height - 40
            c.drawString(40, y, f"Title: {pub.title}")
            y -= 15
            c.drawString(40, y, f"Date: {pub.publication_date.strftime('%Y-%m-%d')}")
            y -= 15
            if pub.abstract:
                abstract = (pub.abstract[:80] + '...') if len(pub.abstract) > 80 else pub.abstract
                c.drawString(40, y, f"Abstract: {abstract}")
                y -= 20
            else:
                y -= 10
        c.save()
        return response
    

class ExportMultiGroupExcel(APIView):
    def get(self, request):
        multi_group_pub_ids = set()
        for pub in Publication.objects.all():
            groups = set(
                pub.authorpublication_set
                .select_related('author__research_group')
                .values_list('author__research_group__name', flat=True)
            )
            if len(groups) > 1:
                multi_group_pub_ids.add(pub.id)

        publications = Publication.objects.filter(id__in=multi_group_pub_ids)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Multi-Group Publications"

        ws.append(["ID", "Title", "Publication Date", "Keywords", "Abstract", "URL", "Source"])

        for pub in publications:
            ws.append([
                pub.id,
                pub.title,
                pub.publication_date.strftime('%Y-%m-%d') if pub.publication_date else "",
                pub.keywords,
                pub.abstract,
                pub.url,
                pub.source
            ])

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 50))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=multi_group_publications.xlsx'
        wb.save(response)
        return response


class ExportMultiGroupPDF(APIView):
    def get(self, request):
        multi_group_pub_ids = set()
        for pub in Publication.objects.all():
            groups = set(
                pub.authorpublication_set
                .select_related('author__research_group')
                .values_list('author__research_group__name', flat=True)
            )
            if len(groups) > 1:
                multi_group_pub_ids.add(pub.id)

        publications = Publication.objects.filter(id__in=multi_group_pub_ids)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=multi_group_publications.pdf'

        c = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        y = height - 40
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, y, "Publications with Multiple Research Groups")
        y -= 30
        c.setFont("Helvetica", 12)

        for pub in publications:
            if y < 80:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica", 12)

            c.drawString(40, y, f"Title: {pub.title}")
            y -= 15
            c.drawString(40, y, f"Date: {pub.publication_date.strftime('%Y-%m-%d') if pub.publication_date else 'N/A'}")
            y -= 15

            if pub.abstract:
                abstract = pub.abstract[:80] + '...' if len(pub.abstract) > 80 else pub.abstract
                c.drawString(40, y, f"Abstract: {abstract}")
                y -= 20
            else:
                y -= 10

        c.save()
        return response
    
class ExportGroupAuthorMultiGroupExcel(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Group Author Multi-Group"

        ws.append(["Group", "Author", "Multi-Group Publication Count"])

        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)

            for author in authors:
                count = 0
                for ap in author.authorpublication_set.all():
                    pub = ap.publication
                    groups = set(
                        pub.authorpublication_set
                        .select_related('author__research_group')
                        .values_list('author__research_group__name', flat=True)
                    )
                    if len(groups) > 1:
                        count += 1

                if count > 0:
                    ws.append([group.name, str(author), count])

        # Auto column width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=group_author_multigroup.xlsx'
        wb.save(response)
        return response


class ExportGroupAuthorMultiGroupPDF(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=group_author_multigroup.pdf'

        c = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        y = height - 40
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, y, "Multi-Group Publications per Author by Group")
        y -= 30
        c.setFont("Helvetica", 12)

        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)
            printed_group = False

            for author in authors:
                count = 0
                for ap in author.authorpublication_set.all():
                    pub = ap.publication
                    groups = set(
                        pub.authorpublication_set
                        .select_related('author__research_group')
                        .values_list('author__research_group__name', flat=True)
                    )
                    if len(groups) > 1:
                        count += 1

                if count > 0:
                    if y < 50:
                        c.showPage()
                        y = height - 40
                        c.setFont("Helvetica", 12)

                    if not printed_group:
                        c.setFont("Helvetica-Bold", 13)
                        c.drawString(40, y, f"Group: {group.name}")
                        y -= 20
                        printed_group = True
                        c.setFont("Helvetica", 12)

                    c.drawString(60, y, f"{author} — {count} publications")
                    y -= 15

        c.save()
        return response
    

class ExportTotalPapersPerGroupExcel(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Papers Per Group"
        ws.append(["Research Group", "Title", "Date", "Keywords", "URL"])

        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)
            pub_ids = set()
            for author in authors:
                pub_ids.update(author.authorpublication_set.values_list('publication_id', flat=True))
            pubs = Publication.objects.filter(id__in=pub_ids)

            for pub in pubs:
                ws.append([
                    group.name,
                    pub.title,
                    pub.publication_date.strftime('%Y-%m-%d') if pub.publication_date else '',
                    pub.keywords or '',
                    pub.url or ''
                ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"papers_per_group.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExportTotalPapersPerGroupPDF(APIView):
    def get(self, request):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        for group in ResearchGroup.objects.all():
            authors = Author.objects.filter(research_group=group)
            pub_ids = set()
            for author in authors:
                pub_ids.update(author.authorpublication_set.values_list('publication_id', flat=True))
            pubs = Publication.objects.filter(id__in=pub_ids)

            if pubs.exists():
                elements.append(Paragraph(f"<b>{group.name}</b>", styles['Heading2']))
                elements.append(Spacer(1, 8))

                for pub in pubs:
                    elements.append(Paragraph(f"<b>Title:</b> {pub.title}", styles['Normal']))
                    if pub.publication_date:
                        elements.append(Paragraph(f"<b>Date:</b> {pub.publication_date.strftime('%Y-%m-%d')}", styles['Normal']))
                    if pub.keywords:
                        elements.append(Paragraph(f"<b>Keywords:</b> {pub.keywords}", styles['Normal']))
                    if pub.url:
                        elements.append(Paragraph(f"<b>URL:</b> {pub.url}", styles['Normal']))
                    elements.append(Spacer(1, 12))

        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"papers_per_group.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response



